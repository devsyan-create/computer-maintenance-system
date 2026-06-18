import openpyxl
import json

wb = openpyxl.load_workbook('new-db.xlsx')
ws = wb.active

print("=== تێگەیشتن لە ستراکچەری Excel ===\n")

# Row 1: Title
print("Row 1 (ناونیشان):", ws['A1'].value)

# Row 2: Headers
print("\nRow 2 (Headers):")
headers = []
for i, cell in enumerate(ws[2], 1):
    if cell.value:
        headers.append((i, cell.value))
        print(f"  Column {i}: {cell.value}")

print("\n=== نموونەی داتا ===\n")

# Sample rows
for row_num in [3, 4, 5, 10, 20]:
    row = list(ws[row_num])
    print(f"\nRow {row_num}:")
    print(f"  شوێن (col 2): {row[1].value}")
    print(f"  شوێن2 (col 3): {row[2].value}")
    print(f"  بەش (col 4): {row[3].value}")
    print(f"  جۆری کەرەستە (col 6): {row[5].value}")
    print(f"  براند (col 7): {row[6].value}")
    print(f"  مۆدێل (col 8): {row[7].value}")

print("\n=== شوێنە جیاوازەکان ===\n")

# Get unique locations from column 2 (شوێن)
locations = set()
for row in ws.iter_rows(min_row=3, values_only=True):
    if row[1]:  # Column 2 (شوێن)
        location = str(row[1]).strip()
        if location:
            locations.add(location)

print(f"ژمارەی شوێنە جیاوازەکان: {len(locations)}")
print("\nلیستی شوێنەکان:")
for loc in sorted(locations):
    print(f"  - {loc}")

print("\n=== کەرەستەکان بەپێی شوێن ===\n")

# Count assets per location
location_counts = {}
for row in ws.iter_rows(min_row=3, values_only=True):
    if row[1] and row[5]:  # Has location and category
        location = str(row[1]).strip()
        if location:
            location_counts[location] = location_counts.get(location, 0) + 1

print("ژمارەی کەرەستە لە هەر شوێنێک:")
for loc in sorted(location_counts.keys()):
    print(f"  {loc}: {location_counts[loc]} کەرەستە")
