import openpyxl
import json

# Load the workbook
wb = openpyxl.load_workbook('new-db.xlsx')

print("Available sheets:", wb.sheetnames)
print()

# Read each sheet
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n=== Sheet: {sheet_name} ===")
    print(f"Rows: {ws.max_row}, Columns: {ws.max_column}")
    
    # Get headers
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)
    
    print("Headers:", headers)
    
    # Get first 5 rows of data
    print("\nSample data:")
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=6, values_only=True), start=2):
        print(f"Row {row_idx}: {row}")
