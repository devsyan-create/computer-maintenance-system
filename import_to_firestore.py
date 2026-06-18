import openpyxl
from datetime import datetime
import json

def generate_serial_number(index):
    """Generate serial number: AST-YYYYMMDD-XXXX"""
    date = datetime.now()
    year = date.year
    month = str(date.month).zfill(2)
    day = str(date.day).zfill(2)
    number = str(index).zfill(4)
    return f"AST-{year}{month}{day}-{number}"

def clean_value(value):
    """Clean cell values"""
    if value is None or value == '':
        return ''
    if isinstance(value, str):
        # Remove formula references
        if value.startswith('='):
            return ''
        return value.strip()
    return str(value)

def normalize_category(category):
    """Normalize category names"""
    category = category.strip()
    category_map = {
        'all in one': 'All In One',
        'desktop': 'Desktop',
        'laptop': 'Laptop',
        'printer': 'Printer',
        'scanner': 'Scanner',
        'moniter': 'Monitor',
        'monitor': 'Monitor',
        'phone': 'Phone',
        'smart phone': 'Smart Phone',
        'tablet': 'Tablet',
        'card printer': 'Card Printer',
        'flash drive': 'Flash Drive',
    }
    
    # Check if it's a brand name being used as category (common mistake in data)
    brands = ['hp', 'dell', 'lenovo', 'canon', 'epson', 'lg', 'apple', 'asus']
    if category.lower() in brands:
        return 'Desktop'  # Default to Desktop for brand-only entries
    
    normalized = category_map.get(category.lower(), category)
    return normalized

# Load the workbook
wb = openpyxl.load_workbook('new-db.xlsx')
ws = wb.active

# Read headers (row 2 has the actual headers)
headers = []
for cell in ws[2]:
    headers.append(clean_value(cell.value))

print("Headers found:", headers[:16])
print()

assets = []
locations_set = set()
categories_set = set()
brands_set = set()

# Read data starting from row 3
for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=1):
    # Skip empty rows
    if not any(row):
        continue
    
    location = clean_value(row[1]) if len(row) > 1 else ''
    location2 = clean_value(row[2]) if len(row) > 2 else ''
    department = clean_value(row[3]) if len(row) > 3 else ''
    category = clean_value(row[5]) if len(row) > 5 else ''
    brand = clean_value(row[6]) if len(row) > 6 else ''
    model = clean_value(row[7]) if len(row) > 7 else ''
    ram = clean_value(row[8]) if len(row) > 8 else ''
    storage = clean_value(row[9]) if len(row) > 9 else ''
    cpu = clean_value(row[10]) if len(row) > 10 else ''
    user = clean_value(row[11]) if len(row) > 11 else ''
    serial = clean_value(row[12]) if len(row) > 12 else ''
    phone = clean_value(row[13]) if len(row) > 13 else ''
    job = clean_value(row[14]) if len(row) > 14 else ''
    notes = clean_value(row[15]) if len(row) > 15 else ''
    
    # Skip if no category
    if not category:
        continue
    
    # Normalize category
    category = normalize_category(category)
    
    # Combine location fields
    full_location = location
    if location2:
        full_location = f"{location} - {location2}"
    if department:
        full_location = f"{full_location} - {department}" if full_location else department
    
    # Add to sets for later creation
    if full_location:
        locations_set.add(full_location)
    if category:
        categories_set.add(category)
    if brand:
        brands_set.add(brand)
    
    # Combine notes
    notes_parts = []
    if notes:
        notes_parts.append(notes)
    if phone:
        notes_parts.append(f"ژمارەی موبایل: {phone}")
    if job:
        notes_parts.append(f"ئیش و کار: {job}")
    
    combined_notes = "\n".join(notes_parts)
    
    # Create asset object matching the website structure
    asset = {
        'serialNumber': serial if serial else generate_serial_number(row_idx),
        'category': category,
        'brand': brand,
        'model': model,
        'cpu': cpu,
        'ram': ram,
        'storage': storage,
        'macAddress': '',
        'location': full_location,
        'user': user,
        'status': 'چالاک',
        'notes': combined_notes,
    }
    
    assets.append(asset)

print(f"✓ Found {len(assets)} assets")
print(f"✓ Found {len(locations_set)} unique locations")
print(f"✓ Found {len(categories_set)} unique categories")
print(f"✓ Found {len(brands_set)} unique brands")
print()

# Save to JSON files
with open('assets_import.json', 'w', encoding='utf-8') as f:
    json.dump(assets, f, ensure_ascii=False, indent=2)

with open('locations_import.json', 'w', encoding='utf-8') as f:
    locations = [{'name': loc} for loc in sorted(locations_set) if loc]
    json.dump(locations, f, ensure_ascii=False, indent=2)

with open('categories_import.json', 'w', encoding='utf-8') as f:
    categories = [{'name': cat} for cat in sorted(categories_set) if cat]
    json.dump(categories, f, ensure_ascii=False, indent=2)

with open('brands_import.json', 'w', encoding='utf-8') as f:
    brands = [{'name': brand} for brand in sorted(brands_set) if brand]
    json.dump(brands, f, ensure_ascii=False, indent=2)

print("✓ JSON files created successfully!")
print()
print("📁 Files:")
print("  → assets_import.json ({} items)".format(len(assets)))
print("  → locations_import.json ({} items)".format(len(locations_set)))
print("  → categories_import.json ({} items)".format(len(categories_set)))
print("  → brands_import.json ({} items)".format(len(brands_set)))
print()
print("📄 Sample asset:")
if assets:
    print(json.dumps(assets[0], ensure_ascii=False, indent=2))
print()
print("📋 Categories found:")
for cat in sorted(categories_set):
    print(f"  - {cat}")
